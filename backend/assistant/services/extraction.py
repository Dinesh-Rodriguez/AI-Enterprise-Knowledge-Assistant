from pathlib import Path

from PIL import Image
from docx import Document as DocxDocument
from openpyxl import load_workbook
from PyPDF2 import PdfReader

try:
    import pytesseract
except Exception:  # pragma: no cover
    pytesseract = None


def extract_text_from_upload(file_path: str) -> tuple[str, list[dict]]:
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix == ".pdf":
        return _extract_pdf(path)
    if suffix in {".docx", ".doc"}:
        return _extract_docx(path), []
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _extract_xlsx(path), []
    if suffix in {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp"}:
        return _extract_image(path)
    return path.read_text(encoding="utf-8", errors="ignore"), []


def _extract_pdf(path: Path):
    reader = PdfReader(str(path))
    pages = []
    text_parts = []
    for index, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        text_parts.append(text)
        pages.append({"page_number": index, "text": text})
    return "\n".join(text_parts), pages


def _extract_docx(path: Path):
    doc = DocxDocument(str(path))
    return "\n".join(paragraph.text for paragraph in doc.paragraphs if paragraph.text.strip())


def _extract_xlsx(path: Path):
    wb = load_workbook(str(path), data_only=True)
    parts = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            values = [str(cell) for cell in row if cell is not None]
            if values:
                parts.append(" | ".join(values))
    return "\n".join(parts)


def _extract_image(path: Path):
    image = Image.open(str(path))
    if pytesseract is None:
        return "", []
    text = pytesseract.image_to_string(image)
    return text, [{"page_number": 1, "text": text}]
